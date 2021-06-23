from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import time
import os
import sqlite3

import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# database connection
conn = sqlite3.connect('./db.sqlite3')
c = conn.cursor()

# get current date
currentDate = time.strftime("%d_%m_%y")

# create a workbook and add a worksheet

c.execute("SELECT * from api_class")


classes = c.fetchall()

for x in classes:

    if(os.path.exists('./reports' + str(x[0]) + '.xlsx')):

        wb = load_workbook(filename="reports" + str(x[0]) + ".xlsx")
        c.execute("SELECT * from api_class")

        sheet = wb['attendance']
        # sheet[ord() + '1']
        for col_index in range(1, 200):
            col = get_column_letter(col_index)

            if sheet['%s%s' % (col, 1)].value is None:
                col2 = get_column_letter(col_index - 1)
                # print sheet.cell('%s%s'% (col2, 1)).value
                if sheet['%s%s' % (col2, 1)].value != currentDate:
                    sheet['%s%s' % (col, 1)] = currentDate
                break

        # saving the file
        wb.save(filename="reports" + str(x[0]) + ".xlsx")

    else:

        c.execute("SELECT * from api_class")

        classes = c.fetchall()
        for x in classes:
            wb = Workbook()
            dest_filename = 'reports'+str(x[0])+'.xlsx'
            c.execute(
                "SELECT s.registration_no, u.username FROM users_user u inner join users_student s on s.user_id = u.id inner join api_enroll a on s.user_id = a.student_id_id where a.class_id_id=%d" % x[0])
            # creating worksheet and giving names to column

            ws1 = wb.active
            ws1.title = "attendance"
            # wbk[f'ws1{x[0]}'].title =f'attendance, {x[0]}'
            ws1.append(('Registration_No', 'Name', currentDate))
            # ws1.append(('', '', ''))

        # entering students information from database
            while True:
                a = c.fetchone()
                if a == None:
                    break
                else:
                    print(a)
                    ws1.append((a[0], a[1]))

        # saving the file
            wb.save(filename=dest_filename)

        # #creating worksheet and giving names to column
        # ws1 = wb.active
        # ws1.title = "attendance"
        # ws1.append(('Registration_No', 'Name', currentDate))
        # ws1.append(('', '', ''))

        # #entering students information from database
        # while True:
        #     a = c.fetchone()
        #     if a == None:
        #         break
        #     else:
        #         ws1.append((a[0], a[1]))

        # #saving the file
        # wb.save(filename = dest_filename)
