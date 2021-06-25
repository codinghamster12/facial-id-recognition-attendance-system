
import cognitive_face as CF
from datetime import date
from .global_variables import personGroupId
import os, urllib
import sqlite3
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, cell, column_index_from_string
import time
import requests
from dotenv import load_dotenv

load_dotenv()

import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
#database connection


currentDate = time.strftime("%d_%m_%y")

def getDateColumn(sheet):
	for i in range(1, sheet.max_column + 1):
		col = get_column_letter(i)
		if sheet['%s%s'% (col,'1')].value == currentDate:
			return col


def openSheet(class_id,token):

    headers = { 'Authorization': 'Token {}'.format(token)}

#get current date

    wb = load_workbook(filename = "reports"+str(class_id)+".xlsx")
    conn = sqlite3.connect('./db.sqlite3')
    c = conn.cursor()

    c.execute("SELECT * from api_class ac where ac.id= %s"%class_id)
    class_ret= c.fetchone()
    
    sheet = wb['attendance']

 



    Key = os.getenv('API_KEY')
    CF.Key.set(Key)

    BASE_URL =  os.getenv('BASE_URL')
      # Replace with your regional Base URL
    CF.BaseUrl.set(BASE_URL)

    # connect = connect = sqlite3.connect("../db.sqlite3")
    # c = connect.cursor()

    attend = [0 for i in range(0,200)]

    currentDir = os.path.dirname(os.path.abspath(__file__)).replace('\\', '/')
    directory = os.path.join(currentDir, 'Cropped_faces')
    for filename in os.listdir(directory):
        if filename.endswith(".jpg"):
            imgurl = os.path.join(directory, filename)
            res = CF.face.detect(imgurl)
            if len(res) != 1:
                print ("No face detected.")
                continue

            faceIds = []
            for face in res:
                faceIds.append(face['faceId'])
            res = CF.face.identify(faceIds, personGroupId)
            for face  in res:
                if not face['candidates']:
                    print ("Unknown")
                else:
                    personId = str(face['candidates'][0]['personId'])
                    
                    c.execute("SELECT * FROM users_student WHERE personID = ?", (personId,))
                    row = c.fetchone()
                    print ('Rowxyz:', row)
                    attend[int(row[0])] += 1
                    print('Attend',attend)
                    print (row[1] + " recognized")

    for row in range(2, sheet.max_row + 1):
        print('Row,', row)
        print(sheet)
        rn = sheet['A%s'% row].value
        print(rn)
        if rn is not None:
            rn = rn[-2:]
            print(rn[-2:])
            if attend[int(rn)] != 0:
                col = getDateColumn(sheet)
                sheet['%s%s' % (col, str(row))] = 1
                c.execute("SELECT user_id from  users_student sc where sc.id= %s"%rn)
                user_id=c.fetchone()
                print(user_id[0])
                myobj={'student_id':user_id[0],'class_id':class_id,'currDate': date.today(),'isEntered':True}
                requests.post('http://127.0.0.1:8000/classes/view-attendance/',myobj,headers=headers)
                
            else:

                col = getDateColumn(sheet)
                sheet['%s%s' % (col, str(row))] = 0
                c.execute("SELECT user_id from  users_student sc where sc.id= %s"%rn)
                user_id=c.fetchone()
                myobj={'student_id':user_id[0],'class_id':class_id,'currDate': date.today(),'isEntered':False}
                requests.post('http://127.0.0.1:8000/classes/view-attendance/',myobj,headers=headers)


    wb.save(filename = "reports" +str(class_id) +".xlsx")



    

    

    import shutil
    folder = './Cropped_faces'
    for the_file in os.listdir(directory):
        file_path = os.path.join(directory, the_file)
        try:
            if os.path.isfile(file_path):
                os.unlink(file_path)
            #elif os.path.isdir(file_path): shutil.rmtree(file_path)
        except Exception as e:
            print(e)


    #currentDir = os.path.dirname(os.path.abspath(__file__))
    #imgurl = urllib.pathname2url(os.path.join(currentDir, "1.jpg"))
    #res = CF.face.detect(imgurl)
    #faceIds = []
    #for face in res:
    #   faceIds.append(face['faceId'])

    #res = CF.face.identify(faceIds,personGroupId)
    # for face in res:
    #     personName = CF.person.get(personGroupId, face['candidates']['personId'])
    #     print personName
    #print res


